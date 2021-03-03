from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class Report:
    def __init__(self):
        self._columns_titles = ['ID zadania', 'Tytuł zadania', 'Opis', 'Data rozpoczęcia', 'Data zakończenia']
        self._report_name_template = '{}.{}.xlsx'
        self._default_column_width = 15

    def generate_report(self, projects, report_location):
        wb = Workbook()
        wb.remove(wb.worksheets[0])  # Remove default worksheet.

        for project in projects:
            ws = wb.create_sheet(project)

            column_index = 0
            for column_title in self._columns_titles:
                column_index += 1
                cell = ws.cell(column=column_index, row=1, value=column_title)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(column_index)].width = self._default_column_width

            ws.column_dimensions[get_column_letter(2)].width = projects[project]['max_summary_len']
            ws.column_dimensions[get_column_letter(3)].width = projects[project]['max_comment_len']

            row_index = 1
            for issue in projects[project]['issues']:
                row_index += 1
                ws.cell(column=1, row=row_index, value=issue)
                ws.cell(column=2, row=row_index, value=projects[project]['issues'][issue]['summary'])
                if 'comment' in projects[project]['issues'][issue]:
                    ws.cell(column=3, row=row_index, value=projects[project]['issues'][issue]['comment'])
                ws.cell(column=4, row=row_index, value=projects[project]['issues'][issue]['min_date'])
                ws.cell(column=5, row=row_index, value=projects[project]['issues'][issue]['max_date'])
        wb.save(filename=report_location)
